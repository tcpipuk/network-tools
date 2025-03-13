"""File handling modules for CLI tools."""

from __future__ import annotations

from collections.abc import Iterable
from csv import DictReader as CSVReader, DictWriter as CSVWriter
from dataclasses import dataclass, field
from json import dump as json_dump, loads as json_loads
from typing import TYPE_CHECKING, Literal

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.workbook import Workbook as OpenPyXLWorkbook

if TYPE_CHECKING:
    from pathlib import Path

    from network_tools.types import JSON_TYPE


@dataclass(slots=True)
class FileReader:
    """Read hosts from a file."""

    path: Path
    type: Literal["csv", "json", "xlsx"]
    data: JSON_TYPE = field(init=False)

    def __post_init__(self) -> None:
        """Initialise the file reader.

        Raises:
            ValueError: If the file type is invalid.
        """
        match self.type:
            case "csv":
                self._read_csv()
            case "json":
                self._read_json()
            case "xlsx":
                self._read_xlsx()
            case _:
                msg = f"Invalid file type: {self.type}"
                raise ValueError(msg)

    def _read_csv(self) -> None:
        """Read data from a CSV file."""
        self.data = list(CSVReader(self.path.read_text().splitlines()))

    def _read_json(self) -> None:
        """Read data from a JSON file."""
        self.data = json_loads(self.path.read_text())

    def _read_xlsx(self) -> None:
        """Read data from an Excel XLSX file."""
        # Load the workbook and get the active worksheet
        worksheet = openpyxl_load_workbook(filename=self.path, data_only=True, read_only=True).active
        # Get the headers from the first row
        headers = [cell.value for cell in next(worksheet.rows)]
        # Extract data into a list of dictionaries (similar to CSV format)
        self.data = []
        for row in list(worksheet.rows)[1:]:  # Skip header row
            self.data.append({header: cell.value for header, cell in zip(headers, row, strict=False)})


@dataclass(slots=True)
class FileWriter:
    """Write data to a file in various formats."""

    path: Path
    type: Literal["csv", "json", "plain", "xlsx"]
    data: dict[str, JSON_TYPE] | list[JSON_TYPE]

    def __post_init__(self) -> None:
        """Initialise the file writer.

        Raises:
            ValueError: If the file type is invalid.
        """
        match self.type:
            case "csv":
                self._write_csv()
            case "json":
                self._write_json()
            case "plain":
                self._write_plain()
            case "xlsx":
                self._write_xlsx()
            case _:
                msg = f"Invalid file type: {self.type}"
                raise ValueError(msg)

    def _write_csv(self) -> None:
        """Write data to a CSV file."""
        writer = CSVWriter(self.path, fieldnames=self.data[0].keys())
        writer.writeheader()
        for row in self.data:
            writer.writerow(row)

    def _write_json(self) -> None:
        """Write data to a JSON file."""
        json_dump(self.data, self.path, indent=2)

    def _write_plain(self) -> None:
        """Write data to a plain text file."""
        # Prepare the content to write based on data type
        content: str
        if isinstance(self.data, list):
            # Join list items with newlines
            content = "\n".join(str(item) for item in self.data)
        elif isinstance(self.data, dict):
            # Format dictionary as "key: value" pairs, one per line
            lines = [f"{key}: {value}" for key, value in self.data.items()]
            content = "\n".join(lines)
        else:
            # Use the data as is for other types, converting to string if needed
            content = str(self.data)
        # Write the prepared content
        self.path.write_text(content)

    def _write_xlsx(self) -> None:
        """Write data to an Excel XLSX file.

        Raises:
            ValueError: If the data is empty.
        """
        # Create a new workbook and select the active worksheet
        workbook = OpenPyXLWorkbook()
        worksheet = workbook.active
        # Raise an error if there is no data to write
        if not self.data:
            msg = "No data to write to file"
            raise ValueError(msg)
        # Handle dictionary - write as key-value pairs
        if isinstance(self.data, dict):
            # Add headers
            worksheet.cell(row=1, column=1, value="Key")
            worksheet.cell(row=1, column=2, value="Value")
            # Write data rows
            for row_idx, (key, value) in enumerate(self.data.items(), 2):
                worksheet.cell(row=row_idx, column=1, value=key)
                # Add multiple cells if the content is iterable
                if isinstance(value, Iterable):
                    [
                        worksheet.cell(row=row_idx, column=col_idx, value=item)
                        for col_idx, item in enumerate(value, 2)
                    ]
                # Otherwise just add a single cell after the key
                else:
                    worksheet.cell(row=row_idx, column=2, value=value)
        # Handle iterable lists of items
        elif isinstance(self.data, Iterable):
            # Use keys for headers if the first item is a dictionary
            if isinstance(self.data[0], dict):
                headers = list(self.data[0].keys())
                [
                    worksheet.cell(row=1, column=col_idx, value=header)
                    for col_idx, header in enumerate(headers, 1)
                ]
                # Write data rows using the same number from the headers
                [
                    worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(key))
                    for row_idx, row_data in enumerate(self.data, 2)
                    for col_idx, key in enumerate(headers, 1)
                ]
            # Handle simple list of values
            else:
                for row_idx, value in enumerate(self.data, 1):
                    [
                        worksheet.cell(row=row_idx, column=col_idx, value=item)
                        for col_idx, item in enumerate(value, 1)
                    ] if isinstance(value, Iterable) else worksheet.cell(row=row_idx, column=1, value=value)
        # Save the workbook
        workbook.save(self.path)
